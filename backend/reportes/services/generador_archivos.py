"""
GeneradorArchivos: Servicio para generar archivos PDF y Excel.
Utiliza reportlab para PDF y openpyxl para Excel.
"""
from typing import Dict, Any, List
from io import BytesIO
from datetime import datetime

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class GeneradorArchivos:
    """
    Servicio para generar archivos de reportes en diferentes formatos.
    """

    def generar_pdf(
        self,
        datos: List[Dict[str, Any]],
        titulo: str,
        subtitulo: str,
        columnas: List[str]
    ) -> BytesIO:
        """
        Genera un archivo PDF con los datos del reporte.
        
        Args:
            datos: Lista de diccionarios con los datos
            titulo: Título principal del reporte
            subtitulo: Subtítulo con período/filtros
            columnas: Lista de nombres de columnas
            
        Returns:
            BytesIO: Buffer con el PDF generado
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Contenedor de elementos
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo para título
        titulo_style = ParagraphStyle(
            'TituloCustom',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=10,
            alignment=TA_CENTER
        )
        
        # Estilo para subtítulo
        subtitulo_style = ParagraphStyle(
            'SubtituloCustom',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Agregar título y subtítulo
        elements.append(Paragraph(titulo, titulo_style))
        elements.append(Paragraph(subtitulo, subtitulo_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Preparar datos para la tabla
        if not datos:
            elements.append(Paragraph('No hay datos para mostrar', styles['Normal']))
        else:
            # Encabezados de tabla
            tabla_data = [columnas]
            
            # Filas de datos
            for item in datos:
                fila = []
                for col in columnas:
                    # Buscar el valor en el diccionario (case-insensitive)
                    valor = None
                    for key in item.keys():
                        if key.lower() == col.lower():
                            valor = item[key]
                            break
                    
                    # Formatear valor
                    if valor is None:
                        fila.append('-')
                    elif isinstance(valor, float):
                        fila.append(f'{valor:,.2f}')
                    else:
                        fila.append(str(valor))
                
                tabla_data.append(fila)
            
            # Crear tabla
            tabla = Table(tabla_data, repeatRows=1)
            
            # Estilo de tabla
            tabla.setStyle(TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Datos
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Bordes
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                
                # Alternar colores de filas
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ]))
            
            elements.append(tabla)
        
        # Agregar pie de página con fecha
        elements.append(Spacer(1, 0.3 * inch))
        pie_style = ParagraphStyle(
            'PieCustom',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        elements.append(
            Paragraph(f'Generado el {fecha_generacion} - SmartSales365', pie_style)
        )
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generar_excel(
        self,
        datos: List[Dict[str, Any]],
        titulo: str,
        subtitulo: str,
        columnas: List[str]
    ) -> BytesIO:
        """
        Genera un archivo Excel con los datos del reporte.
        
        Args:
            datos: Lista de diccionarios con los datos
            titulo: Título principal del reporte
            subtitulo: Subtítulo con período/filtros
            columnas: Lista de nombres de columnas
            
        Returns:
            BytesIO: Buffer con el Excel generado
        """
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte'
        
        # Estilos
        titulo_font = Font(name='Arial', size=14, bold=True, color='1E40AF')
        subtitulo_font = Font(name='Arial', size=10, color='6B7280')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título (fila 1)
        ws.merge_cells('A1:' + chr(64 + len(columnas)) + '1')
        cell_titulo = ws['A1']
        cell_titulo.value = titulo
        cell_titulo.font = titulo_font
        cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtítulo (fila 2)
        ws.merge_cells('A2:' + chr(64 + len(columnas)) + '2')
        cell_subtitulo = ws['A2']
        cell_subtitulo.value = subtitulo
        cell_subtitulo.font = subtitulo_font
        cell_subtitulo.alignment = Alignment(horizontal='center', vertical='center')
        
        # Espacio
        ws.row_dimensions[3].height = 10
        
        # Encabezados (fila 4)
        fila_inicio_datos = 4
        for col_idx, col_name in enumerate(columnas, start=1):
            cell = ws.cell(row=fila_inicio_datos, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        # Datos (desde fila 5)
        fila_actual = fila_inicio_datos + 1
        
        for item in datos:
            for col_idx, col_name in enumerate(columnas, start=1):
                # Buscar valor (case-insensitive)
                valor = None
                for key in item.keys():
                    if key.lower() == col_name.lower():
                        valor = item[key]
                        break
                
                cell = ws.cell(row=fila_actual, column=col_idx)
                cell.value = valor if valor is not None else '-'
                cell.font = data_font
                cell.border = border_thin
                
                # Alineación según tipo
                if isinstance(valor, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(valor, float):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left')
            
            fila_actual += 1
        
        # Ajustar anchos de columna
        for col_idx in range(1, len(columnas) + 1):
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = 18
        
        # Pie de página
        fila_pie = fila_actual + 2
        ws.merge_cells(f'A{fila_pie}:' + chr(64 + len(columnas)) + str(fila_pie))
        cell_pie = ws[f'A{fila_pie}']
        fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        cell_pie.value = f'Generado el {fecha_generacion} - SmartSales365'
        cell_pie.font = Font(name='Arial', size=8, color='6B7280', italic=True)
        cell_pie.alignment = Alignment(horizontal='center')
        
        # Guardar en buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def determinar_columnas(
        self,
        tipo_reporte: str,
        agrupacion: str
    ) -> List[str]:
        """
        Determina las columnas apropiadas según el tipo de reporte.
        
        Args:
            tipo_reporte: 'ventas' | 'productos' | 'clientes' | 'ingresos'
            agrupacion: 'producto' | 'cliente' | 'categoria' | 'fecha' | 'ninguno'
            
        Returns:
            List[str]: Lista de nombres de columnas
        """
        if tipo_reporte == 'ventas':
            if agrupacion == 'producto':
                return ['Producto', 'Total Vendido', 'Cantidad Ventas', 'Ticket Promedio']
            elif agrupacion == 'cliente':
                return ['Cliente', 'Email', 'Total Gastado', 'Cantidad Compras', 'Ticket Promedio']
            elif agrupacion == 'fecha':
                return ['Fecha', 'Total Vendido', 'Cantidad Ventas', 'Ticket Promedio']
            else:
                return ['Total Vendido', 'Cantidad Ventas', 'Ticket Promedio']
        
        elif tipo_reporte == 'productos':
            return ['nombre', 'categoria', 'precio', 'stock', 'activo']
        
        elif tipo_reporte == 'clientes':
            return ['Cliente', 'Email', 'Total Gastado', 'Cantidad Compras']
        
        elif tipo_reporte == 'ingresos':
            return ['Total Ingresos', 'Cantidad Transacciones', 'Ticket Promedio']
        
        return ['Dato', 'Valor']
