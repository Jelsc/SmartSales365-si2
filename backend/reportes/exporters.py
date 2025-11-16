"""
Exportadores de reportes a diferentes formatos (PDF, Excel)
"""
from io import BytesIO
from typing import Dict
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class PDFExporter:
    """Exportador a PDF usando ReportLab"""
    
    def generar(self, datos_reporte: Dict) -> BytesIO:
        """
        Generar PDF del reporte
        
        Args:
            datos_reporte: Diccionario con datos, columnas, título, etc.
        
        Returns:
            BytesIO con el contenido del PDF
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elementos = []
        styles = getSampleStyleSheet()
        
        # Estilo para título
        style_titulo = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            alignment=1  # Centrado
        )
        
        # Estilo para subtítulo
        style_subtitulo = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=1
        )
        
        # Agregar título
        titulo = Paragraph(datos_reporte['titulo'], style_titulo)
        elementos.append(titulo)
        
        # Agregar subtítulo
        if datos_reporte.get('subtitulo'):
            subtitulo = Paragraph(datos_reporte['subtitulo'], style_subtitulo)
            elementos.append(subtitulo)
        
        # Agregar información de generación
        info = Paragraph(
            f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')} | "
            f"Total de registros: {datos_reporte['total_registros']}",
            styles['Normal']
        )
        elementos.append(info)
        elementos.append(Spacer(1, 0.3 * inch))
        
        # Crear tabla de datos
        if datos_reporte['datos']:
            # Preparar datos para la tabla
            columnas = datos_reporte['columnas']
            datos_tabla = [columnas]  # Primera fila: encabezados
            
            # Agregar filas de datos
            for registro in datos_reporte['datos']:
                fila = []
                for col in columnas:
                    # Mapear nombres de columnas a claves del diccionario
                    key = self._normalizar_key(col)
                    valor = registro.get(key, '')
                    
                    # Formatear valores
                    if isinstance(valor, float):
                        if 'total' in key or 'monto' in key or 'precio' in key or 'vendido' in key:
                            valor = f"Bs {valor:,.2f}"
                        else:
                            valor = f"{valor:,.2f}"
                    
                    fila.append(str(valor))
                datos_tabla.append(fila)
            
            # Crear tabla
            tabla = Table(datos_tabla)
            
            # Estilo de la tabla
            tabla.setStyle(TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Datos
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Bordes
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e40af')),
                
                # Alternar colores de fila
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ]))
            
            elementos.append(tabla)
        else:
            elementos.append(Paragraph("No hay datos para mostrar", styles['Normal']))
        
        # Construir PDF
        doc.build(elementos)
        buffer.seek(0)
        return buffer
    
    def _normalizar_key(self, columna: str) -> str:
        """Convertir nombre de columna a key del diccionario"""
        # "Nombre del Cliente" -> "nombre_cliente"
        # "Cantidad de Pedidos" -> "cantidad_pedidos"
        normalized = columna.lower()
        # Remover palabras comunes que no están en las claves
        normalized = normalized.replace(' de ', '_').replace(' del ', '_').replace(' la ', '_')
        # Reemplazar espacios restantes con guiones bajos
        normalized = normalized.replace(' ', '_')
        # Remover acentos
        normalized = normalized.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
        return normalized
    
    def generar_multiple(self, reportes: list) -> BytesIO:
        """
        Generar PDF con múltiples reportes en el mismo documento
        
        Args:
            reportes: Lista de diccionarios con datos de reportes
        
        Returns:
            BytesIO con el contenido del PDF
        """
        buffer = BytesIO()
        
        # Configuración del documento (usando el tamaño de página del primer reporte)
        tamaño_pagina = letter
        doc = SimpleDocTemplate(
            buffer,
            pagesize=tamaño_pagina,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elementos = []
        styles = getSampleStyleSheet()
        
        # Agregar cada reporte
        for idx, datos_reporte in enumerate(reportes):
            # Si no es el primer reporte, agregar salto de página
            if idx > 0:
                elementos.append(PageBreak())
            
            # Encabezado del documento
            titulo_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=12,
                alignment=1  # Centrado
            )
            
            titulo = Paragraph(datos_reporte['titulo'], titulo_style)
            elementos.append(titulo)
            
            # Subtítulo
            if datos_reporte.get('subtitulo'):
                subtitulo_style = ParagraphStyle(
                    'CustomSubtitle',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.grey,
                    spaceAfter=8,
                    alignment=1
                )
                subtitulo = Paragraph(datos_reporte['subtitulo'], subtitulo_style)
                elementos.append(subtitulo)
            
            # Información adicional
            info_style = ParagraphStyle(
                'CustomInfo',
                parent=styles['Normal'],
                fontSize=9,
                spaceAfter=20,
                alignment=1
            )
            info = Paragraph(
                f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')} | Total de registros: {datos_reporte['total_registros']}",
                info_style
            )
            elementos.append(info)
            
            # Indicador de reporte múltiple
            if len(reportes) > 1:
                indicador_style = ParagraphStyle(
                    'Indicador',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#1e40af'),
                    spaceAfter=10,
                    alignment=1,
                    fontName='Helvetica-Bold'
                )
                indicador = Paragraph(f"Reporte {idx + 1} de {len(reportes)}", indicador_style)
                elementos.append(indicador)
            
            # Tabla de datos
            if datos_reporte['datos']:
                # Encabezados
                datos_tabla = [datos_reporte['columnas']]
                
                # Filas de datos
                for registro in datos_reporte['datos']:
                    fila = []
                    for columna in datos_reporte['columnas']:
                        key = self._normalizar_key(columna)
                        valor = registro.get(key, '')
                        
                        # Formatear valores monetarios
                        if isinstance(valor, (int, float)):
                            if 'total' in key or 'monto' in key or 'precio' in key or 'vendido' in key:
                                valor = f"Bs {valor:,.2f}"
                            else:
                                valor = f"{valor:,.2f}"
                        
                        fila.append(str(valor))
                    
                    datos_tabla.append(fila)
                
                # Crear tabla
                tabla = Table(datos_tabla)
                
                # Estilo de la tabla
                tabla.setStyle(TableStyle([
                    # Encabezado
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
                    # Datos
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                    
                    # Bordes
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e40af')),
                    
                    # Alternar colores de fila
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                ]))
                
                elementos.append(tabla)
            else:
                elementos.append(Paragraph("No hay datos para mostrar", styles['Normal']))
        
        # Construir PDF
        doc.build(elementos)
        buffer.seek(0)
        return buffer


class ExcelExporter:
    """Exportador a Excel usando openpyxl"""
    
    def generar(self, datos_reporte: Dict) -> BytesIO:
        """
        Generar Excel del reporte
        
        Args:
            datos_reporte: Diccionario con datos, columnas, título, etc.
        
        Returns:
            BytesIO con el contenido del Excel
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte"
        
        # Estilos
        titulo_font = Font(size=16, bold=True, color="1e40af")
        subtitulo_font = Font(size=10, color="666666")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        fila_actual = 1
        
        # Título
        col_count = len(datos_reporte['columnas'])
        sheet.merge_cells(f'A{fila_actual}:{get_column_letter(col_count)}{fila_actual}')
        cell_titulo = sheet[f'A{fila_actual}']
        cell_titulo.value = datos_reporte['titulo']
        cell_titulo.font = titulo_font
        cell_titulo.alignment = Alignment(horizontal='center')
        fila_actual += 1
        
        # Subtítulo
        if datos_reporte.get('subtitulo'):
            sheet.merge_cells(f'A{fila_actual}:{get_column_letter(col_count)}{fila_actual}')
            cell_subtitulo = sheet[f'A{fila_actual}']
            cell_subtitulo.value = datos_reporte['subtitulo']
            cell_subtitulo.font = subtitulo_font
            cell_subtitulo.alignment = Alignment(horizontal='center')
            fila_actual += 1
        
        # Info
        sheet.merge_cells(f'A{fila_actual}:{get_column_letter(col_count)}{fila_actual}')
        cell_info = sheet[f'A{fila_actual}']
        cell_info.value = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')} | Total de registros: {datos_reporte['total_registros']}"
        cell_info.font = Font(size=9)
        cell_info.alignment = Alignment(horizontal='center')
        fila_actual += 2  # Espacio
        
        # Encabezados
        for col_idx, columna in enumerate(datos_reporte['columnas'], start=1):
            cell = sheet.cell(row=fila_actual, column=col_idx)
            cell.value = columna
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        fila_actual += 1
        
        # Datos
        for registro in datos_reporte['datos']:
            for col_idx, columna in enumerate(datos_reporte['columnas'], start=1):
                key = self._normalizar_key(columna)
                valor = registro.get(key, '')
                
                cell = sheet.cell(row=fila_actual, column=col_idx)
                
                # Establecer valor (mantener tipos para Excel)
                if isinstance(valor, str) and valor.startswith('Bs '):
                    # Convertir "Bs 1234.56" a número
                    try:
                        cell.value = float(valor.replace('Bs ', '').replace(',', ''))
                        cell.number_format = '"Bs "#,##0.00'
                    except:
                        cell.value = valor
                elif isinstance(valor, (int, float)):
                    cell.value = valor
                    if 'total' in key or 'monto' in key or 'precio' in key or 'vendido' in key:
                        cell.number_format = '"Bs "#,##0.00'
                    else:
                        cell.number_format = '#,##0.00'
                else:
                    cell.value = str(valor)
                
                cell.border = border
                
                # Alternar colores
                if fila_actual % 2 == 0:
                    cell.fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
            
            fila_actual += 1
        
        # Ajustar anchos de columna
        for col_idx in range(1, col_count + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            
            for cell in sheet[col_letter]:
                # Saltar celdas combinadas
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[col_letter].width = adjusted_width
        
        # Guardar en BytesIO
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generar_multiple(self, reportes: list) -> BytesIO:
        """
        Generar Excel con múltiples reportes en hojas separadas
        
        Args:
            reportes: Lista de diccionarios con datos de reportes
        
        Returns:
            BytesIO con el contenido del Excel
        """
        workbook = openpyxl.Workbook()
        # Eliminar la hoja por defecto
        workbook.remove(workbook.active)
        
        # Estilos comunes
        titulo_font = Font(size=16, bold=True, color="1e40af")
        subtitulo_font = Font(size=10, color="666666")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Crear una hoja por cada reporte
        for idx, datos_reporte in enumerate(reportes):
            # Nombre de la hoja (limitado a 31 caracteres)
            titulo_hoja = datos_reporte.get('titulo', f'Reporte {idx + 1}')[:31]
            sheet = workbook.create_sheet(title=titulo_hoja)
            
            col_count = len(datos_reporte['columnas'])
            fila_actual = 1
            
            # Indicador de reporte múltiple
            if len(reportes) > 1:
                sheet.merge_cells(f'A{fila_actual}:{get_column_letter(col_count)}{fila_actual}')
                cell_indicador = sheet[f'A{fila_actual}']
                cell_indicador.value = f"Reporte {idx + 1} de {len(reportes)}"
                cell_indicador.font = Font(size=12, bold=True, color="1e40af")
                cell_indicador.alignment = Alignment(horizontal='center')
                fila_actual += 1
            
            # Título
            sheet.merge_cells(f'A{fila_actual}:{get_column_letter(col_count)}{fila_actual}')
            cell_titulo = sheet[f'A{fila_actual}']
            cell_titulo.value = datos_reporte['titulo']
            cell_titulo.font = titulo_font
            cell_titulo.alignment = Alignment(horizontal='center')
            fila_actual += 1
            
            # Subtítulo
            if datos_reporte.get('subtitulo'):
                sheet.merge_cells(f'A{fila_actual}:{get_column_letter(col_count)}{fila_actual}')
                cell_subtitulo = sheet[f'A{fila_actual}']
                cell_subtitulo.value = datos_reporte['subtitulo']
                cell_subtitulo.font = subtitulo_font
                cell_subtitulo.alignment = Alignment(horizontal='center')
                fila_actual += 1
            
            # Info
            sheet.merge_cells(f'A{fila_actual}:{get_column_letter(col_count)}{fila_actual}')
            cell_info = sheet[f'A{fila_actual}']
            cell_info.value = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')} | Total de registros: {datos_reporte['total_registros']}"
            cell_info.font = Font(size=9)
            cell_info.alignment = Alignment(horizontal='center')
            fila_actual += 2  # Espacio
            
            # Encabezados
            for col_idx, columna in enumerate(datos_reporte['columnas'], start=1):
                cell = sheet.cell(row=fila_actual, column=col_idx)
                cell.value = columna
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            fila_actual += 1
            
            # Datos
            for registro in datos_reporte['datos']:
                for col_idx, columna in enumerate(datos_reporte['columnas'], start=1):
                    key = self._normalizar_key(columna)
                    valor = registro.get(key, '')
                    
                    cell = sheet.cell(row=fila_actual, column=col_idx)
                    
                    # Establecer valor (mantener tipos para Excel)
                    if isinstance(valor, str) and valor.startswith('Bs '):
                        # Convertir "Bs 1234.56" a número
                        try:
                            cell.value = float(valor.replace('Bs ', '').replace(',', ''))
                            cell.number_format = '"Bs "#,##0.00'
                        except:
                            cell.value = valor
                    elif isinstance(valor, (int, float)):
                        cell.value = valor
                        if 'total' in key or 'monto' in key or 'precio' in key or 'vendido' in key:
                            cell.number_format = '"Bs "#,##0.00'
                        else:
                            cell.number_format = '#,##0.00'
                    else:
                        cell.value = str(valor)
                    
                    cell.border = border
                    
                    # Alternar colores
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
                
                fila_actual += 1
            
            # Ajustar anchos de columna
            for col_idx in range(1, col_count + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                for cell in sheet[col_letter]:
                    # Saltar celdas combinadas
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[col_letter].width = adjusted_width
        
        # Guardar en BytesIO
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _normalizar_key(self, columna: str) -> str:
        """Convertir nombre de columna a key del diccionario"""
        # "Nombre del Cliente" -> "nombre_cliente"
        # "Cantidad de Pedidos" -> "cantidad_pedidos"
        normalized = columna.lower()
        # Remover palabras comunes que no están en las claves
        normalized = normalized.replace(' de ', '_').replace(' del ', '_').replace(' la ', '_')
        # Reemplazar espacios restantes con guiones bajos
        normalized = normalized.replace(' ', '_')
        # Remover acentos
        normalized = normalized.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
        return normalized
